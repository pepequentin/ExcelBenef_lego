import pandas as pd
import requests
from bs4 import BeautifulSoup
import re
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from datetime import date, timedelta
import openpyxl
from openpyxl.styles import PatternFill
from tqdm import tqdm


##
#   Utilisation du script avec le fichier 'Achat_lego.xlsx'
##
def check_prices(file_path):
    df = pd.read_excel(file_path)
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    vente_total = 0
    num_rows = len(df)

    # Créez une instance tqdm
    progress_bar = tqdm(total=num_rows, desc="Processing", unit="rows")

    for index, row in df.iterrows():
        lien = row[1]
        prix_achat = row[6]
        nb_exemplaires = row[12]

        if pd.notna(lien) and isinstance(lien, str) and pd.notna(prix_achat) and nb_exemplaires > 0:
            user_agent = 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/113.0.0.0 Safari/537.36'
            headers = {'User-Agent': user_agent}
            response = requests.get(lien, headers=headers)
            html_source_code = response.text
            soup = BeautifulSoup(html_source_code, "html.parser")
            html_span = soup.find_all('span', {'class': 'oopStage-priceRangePrice'})
            pattern = re.compile(r'<span class="oopStage-conditionButton-wrapper-text-price-prefix">\s*\(non disponible\)\s*</span>')
            html_source_code = response.text
            html_span_for_no_news = pattern.findall(html_source_code)
            prix_pattern = re.compile(r'\d+,\d+')
            prix_achat_par_exemplaire = prix_achat / nb_exemplaires

            # Condition pour les lego qui non pas de prix idealo
            if not html_span:
                prix_trouve = row[8]
                if prix_trouve:
                    prix = prix_trouve
                    if prix:
                        if prix_achat != 0:
                            pourcentage_benef = ((prix - prix_achat_par_exemplaire) / prix_achat_par_exemplaire) * 100
                        else:
                            pourcentage_benef = ((prix - prix_achat_par_exemplaire) / 1) * 100


                        # Prix mul by the number of product in a temp var
                        tmp_price = prix * nb_exemplaires
                        vente_total += tmp_price
                        df.at[index, 'Prix actuel idéalo'] = f"{prix:.2f}"
                        df.at[index, 'Bénéfice potentiel'] = f"{pourcentage_benef * nb_exemplaires:.2f}"

                        # Appliquer le formatage des couleurs en fonction des valeurs de 'Bénéfice potentiel'
                        if pd.notna(row['Bénéfice potentiel']):
                            pourcentage_benef = float(row['Bénéfice potentiel'][:-1])
                            if pourcentage_benef > 0:
                                ws.cell(row=index + 2, column=11).fill = openpyxl.styles.PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
                            elif pourcentage_benef < 0:
                                ws.cell(row=index + 2, column=11).fill = openpyxl.styles.PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                            else:
                                ws.cell(row=index + 2, column=11).fill = openpyxl.styles.PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type="solid")
            else:
                for span in html_span:
                    if html_span_for_no_news:
                        prix_trouve = prix_pattern.search(span.text)
                        if prix_trouve:
                            prix = float(prix_trouve.group().replace(",", "."))
                            if prix:
                                prix *= 4
                                if prix_achat != 0:
                                    pourcentage_benef = ((prix - prix_achat_par_exemplaire) / prix_achat_par_exemplaire) * 100
                                else:
                                    pourcentage_benef = ((prix - prix_achat_par_exemplaire) / 1) * 100


                                # Prix mul by the number of product in a temp var
                                tmp_price = prix * nb_exemplaires
                                vente_total += tmp_price
                                df.at[index, 'Prix actuel idéalo'] = f"{prix:.2f}"
                                df.at[index, 'Bénéfice potentiel'] = f"{pourcentage_benef * nb_exemplaires:.2f}"

                                # Appliquer le formatage des couleurs en fonction des valeurs de 'Bénéfice potentiel'
                                if pd.notna(row['Bénéfice potentiel']):
                                    pourcentage_benef = float(row['Bénéfice potentiel'][:-1])
                                    if pourcentage_benef > 0:
                                        ws.cell(row=index + 2, column=11).fill = openpyxl.styles.PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
                                    elif pourcentage_benef < 0:
                                        ws.cell(row=index + 2, column=11).fill = openpyxl.styles.PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                                    else:
                                        ws.cell(row=index + 2, column=11).fill = openpyxl.styles.PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type="solid")

                    else:
                        prix_trouve = prix_pattern.search(span.text)
                        if prix_trouve:
                            prix = float(prix_trouve.group().replace(",", "."))
                            if prix:
                                if prix_achat != 0:
                                    pourcentage_benef = ((prix - prix_achat_par_exemplaire) / prix_achat_par_exemplaire) * 100
                                else:
                                    pourcentage_benef = ((prix - prix_achat_par_exemplaire) / 1) * 100


                                # Prix mul by the number of product in a temp var
                                tmp_price = prix * nb_exemplaires
                                vente_total += tmp_price
                                df.at[index, 'Prix actuel idéalo'] = f"{prix:.2f}"
                                df.at[index, 'Bénéfice potentiel'] = f"{pourcentage_benef * nb_exemplaires:.2f}"

                                # Appliquer le formatage des couleurs en fonction des valeurs de 'Bénéfice potentiel'
                                if pd.notna(row['Bénéfice potentiel']):
                                    pourcentage_benef = float(row['Bénéfice potentiel'][:-1])
                                    if pourcentage_benef > 0:
                                        ws.cell(row=index + 2, column=11).fill = openpyxl.styles.PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
                                    elif pourcentage_benef < 0:
                                        ws.cell(row=index + 2, column=11).fill = openpyxl.styles.PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                                    else:
                                        ws.cell(row=index + 2, column=11).fill = openpyxl.styles.PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type="solid")
            
            # Mettez à jour la barre de progression
            progress_bar.update(1)

    # Mettez à jour la barre de progression
    progress_bar.update(2)

    # Fermez la barre de progression
    progress_bar.close()
    
    final_output_file = 'Achat_lego_updated.xlsx'
    wb.save(final_output_file)
    df.to_excel(final_output_file, index=False)

    # Charger le classeur pour appliquer le formatage des couleurs
    new_wb = openpyxl.load_workbook(final_output_file)
    ws = new_wb.active

    # Appliquer le formatage des couleurs en fonction des valeurs de 'Bénéfice potentiel'
    for index, row in df.iterrows():
        if pd.notna(row['Bénéfice potentiel']):
            pourcentage_benef = float(row['Bénéfice potentiel'][:-1])
            if pourcentage_benef > 0:
                ws.cell(row=index + 2, column=11).fill = openpyxl.styles.PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
            elif pourcentage_benef < 0:
                ws.cell(row=index + 2, column=11).fill = openpyxl.styles.PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
            else:
                ws.cell(row=index + 2, column=11).fill = openpyxl.styles.PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type="solid")

    # Calculer le coût total
    cout_total = df['Prix d\'achat'].sum()
    ws.append([])
    ws.append([None, None, "Cout total", cout_total])

    # Calculer la vente total
    ws.append([None, None, "Vente total", vente_total])

    # Calculer le potentiel bénéficiaire
    ws.append([None, None, "Potentiel benef", vente_total - cout_total])


    final_output_file = 'Achat_lego_updated.xlsx'
    new_wb.save(final_output_file)

    # Ouvrir les fichiers Excel
    wb1 = openpyxl.load_workbook(file_path)
    wb2 = openpyxl.load_workbook('Achat_lego_updated.xlsx')

    # Sélectionner les feuilles actives des deux fichiers
    ws1 = wb1.active
    ws2 = wb2.active

    # Spécifier les colonnes à copier
    col_to_copy = 'H'

    for row_num, (row1, row2) in enumerate(zip(ws1.iter_rows(min_row=2, values_only=True), ws2.iter_rows(min_row=2, values_only=True)), start=2):
        color1 = ws1[f'{col_to_copy}{row_num}'].fill.start_color
        ws2[f'{col_to_copy}{row_num}'].fill = openpyxl.styles.PatternFill(start_color=color1, end_color=color1, fill_type="solid")
   
    col_to_copy = 'R'

    for row_num, (row1, row2) in enumerate(zip(ws1.iter_rows(min_row=2, values_only=True), ws2.iter_rows(min_row=2, max_row=5, values_only=True)), start=2):
        color1 = ws1[f'{col_to_copy}{row_num}'].fill.start_color
        ws2[f'{col_to_copy}{row_num}'].fill = openpyxl.styles.PatternFill(start_color=color1, end_color=color1, fill_type="solid")

    # Copier la largeur de colonne du fichier1 sur le fichier2
    for column in ws1.column_dimensions:
        ws2.column_dimensions[column].width = ws1.column_dimensions[column].width

    # Sauvegarder les modifications dans le fichier2
    wb2.save('Achat_lego_updated.xlsx')
    print("Données mises à jour et sauvegardées dans", final_output_file)


if __name__ == "__main__":
    # Utilisation du script avec le fichier 'Achat_lego.xlsx'
    check_prices('../excel_idealo/Achat_lego.xlsx')